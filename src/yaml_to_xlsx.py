import collections
import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from src import utils
from src import myclass
from src import config

# Define the fill colors for alternating rows
even_fill = PatternFill(start_color='C6EFCE',
                        end_color='C6EFCE',
                        fill_type='solid')
odd_fill = PatternFill(start_color='FFFFFF',
                       end_color='FFFFFF',
                       fill_type='solid')
header_fill = PatternFill(start_color='4ecd5f',
                          end_color='FFFFFF',
                          fill_type='solid')


def convert_story_tuple_to_sheet(ws, story_tuples, idx, col):
    domain_to_tuples = collections.defaultdict(list)
    for tup in story_tuples:
        domain_to_tuples[tup['domain']].append(tup)

    for domain, tuples in domain_to_tuples.items():
        tuples = list(map(utils.parse_tuple, tuples))
        ws.cell(row=idx, column=col).value = domain
        ws.cell(row=idx, column=col).fill = even_fill
        idx += 1
        for tup in tuples:
            ws.cell(row=idx, column=col).value = tup
            idx += 1
    return idx


def convert_yaml_to_xlsx(task_id, raw_yaml_path, output_path):
    example = utils.read_yaml(raw_yaml_path)['example']
    example = myclass.Example.from_dict(example)

    turns = [turn.to_dict() for turn in example.history_with_deleted_turns]
    df = pd.DataFrame.from_dict(turns)
    df['annotations'] = df['annotations'].apply(
        lambda x: '\n'.join(list(map(utils.parse_tuple, x))))
    df.insert(0, 'Feedback', '')

    task_table = utils.read_yaml('./data/task_table.yaml')
    task = task_table[task_id]
    flow_path = f"./data/flows/{task['accident_location']}/{task_id}.yaml"
    flow = utils.read_yaml(flow_path)

    rows = dataframe_to_rows(df, index=False)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('Dialog')

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    cols = ['A', 'E', 'G', 'H']

    for col in cols:
        ws.column_dimensions[col].width = 35
        for row in ws[col]:
            row.alignment = Alignment(wrap_text=True)

    # Apply alternating colors to every other row
    for row in ws.iter_rows():
        if row[0].row == 1:
            for cell in row:
                cell.fill = header_fill
        elif row[0].row % 2 == 0:
            for cell in row:
                cell.fill = even_fill
        else:
            for cell in row:
                cell.fill = odd_fill

    ws = wb.create_sheet('Caller Tuples')
    ws.column_dimensions['A'].width = 60
    idx = 1
    for macro in [
            '[Adjuster_TUPLES]', '[ContactInfo_TUPLES]',
            '[DriverActions_TUPLES]', '[Evidences_TUPLES]', '[CarInfo_TUPLES]',
            '[TrafficEnvironment_TUPLES]', '[InjuryDetails_TUPLES]'
    ]:
        if macro not in flow:
            continue
        tuples = list(map(utils.parse_tuple, flow[macro]))
        ws.cell(row=idx, column=1).value = macro
        ws.cell(row=idx, column=1).fill = even_fill
        idx += 1
        for tup in tuples:
            ws.cell(row=idx, column=1).value = tup
            idx += 1

    if 'story_caller_tuples' in flow:
        ws = wb.create_sheet('Tuples For Story')
        idx = 1
        ws.column_dimensions['A'].width = 60
        ws.cell(row=idx, column=1).value = 'caller tuples'
        ws.cell(row=idx, column=1).fill = header_fill
        idx += 1
        idx = convert_story_tuple_to_sheet(ws, flow['story_caller_tuples'],
                                           idx, 1)
        idx += 1

        idx = 1
        ws.column_dimensions['B'].width = 60
        ws.cell(row=idx, column=2).value = 'other driver tuples'
        ws.cell(row=idx, column=2).fill = header_fill
        idx += 1
        convert_story_tuple_to_sheet(ws, flow['story_other_driver_tuples'],
                                     idx, 2)

        flow['personality'] = [
            flow['user_personality'], flow['agent_personality']
        ]

        ws = wb.create_sheet('Prompt')
        ws.column_dimensions['A'].width = 200
        idx = 1
        for key in [
                'story_summaries', 'story', 'information_summaries',
                'personality', 'step_summaries', 'instructions'
        ]:
            summaries = flow[key]
            ws.cell(row=idx, column=1).value = key
            ws.cell(row=idx, column=1).fill = even_fill
            idx += 1
            for summary in summaries:
                ws.cell(row=idx, column=1).value = summary
                ws.cell(row=idx,
                        column=1).alignment = Alignment(wrap_text=True)
                idx += 1
            idx += 1

    wb.save(output_path)
